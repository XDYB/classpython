# -*- coding: utf-8 -*-
# @Author  : zgh
# @Email   : 849080458@qq.com
# @File    : writeExcel.py
# @Software: PyCharm


from openpyxl import Workbook,load_workbook
from interface_auto_cases_for_ZNKF.conf import Allpath
from interface_auto_cases_for_ZNKF.public.logger import Log

logger=Log('auto_cases',Allpath.log_path)


class writeExcel:
    def __init__(self,file,sheet):
        self.file=file
        self.sheet=sheet

    def write_Excel(self,i,dict):

        try:
            wb_new = load_workbook(self.file)
            sheet = wb_new[self.sheet]
            sheet.cell(i, 8).value = dict['code']
            # sheet.cell(i,9).value = dcit['sql_result']
            sheet.cell(i, 10).value = dict['result']
            if 'data' in dict.keys():
                print(dict['data'])
                sheet.cell(i, 11).value = str(dict['data'])
            wb_new.save(self.file)
            logger.info('执行写入excel成功！')
        except Exception as e:
            logger.info('数据写入失败！%s'%e)
            raise e

        '''
        if os.path.exists(self.file):
            print('文件存在，不需要创建！')
            we=load_workbook(self.file)
            if self.sheet in we.get_sheet_names():
                print('表单存在，不需要创建！')
            else:
                print('表单不存在，已重新创建！')
                we.create_sheet(self.sheet)
                we.save(self.file)
        else:
            print('文件不存在，已重新创建！')
            we=Workbook()
            we.create_sheet(self.sheet)
            we.save(self.sheet)

        we_new=load_workbook(self.file)
        sheet=we_new.get_sheet_by_name(self.sheet)

        #先写头部
        #for i in range(len(header)):
            #sheet.cell(row=1,column=i+1).value=header[i]

        #写数据
        #for i in range(len(data)):
            #for ii in range(len(data[i])):
                #sheet.cell(row=i+2,column=ii+1).value=data[i][ii]'''






if __name__ == '__main__':
    header = ['学号', '姓名', '性别', '班级']
    data = [['001', 'sunny', 'F', 'python5'], ['002', '小胖雨', 'M', 'python5']]
    print(writeExcel(Allpath.test_data_path,'Sheet1').write_Excel(1,{'code':'zgh','sql_result':000,'result':333,'msg':555}))